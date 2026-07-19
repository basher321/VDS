"""Parse the Mushak-6.6 'Summary' workbook into staged invoice rows.

Expected columns (header row on the Summary sheet), matched case-insensitively:
  Supplier's Name | Supplier's BIN | Tax Deposit Serial No. | Tax Deposit Date |
  Concerned Invoice number | Invoice/Challan/Bill date | Deducted VAT | VAT Rate |
  Notes (service category)
The base value is DERIVED (VAT / rate), so it need not be present in the file.
"""
from datetime import date, datetime
import openpyxl

# maps a normalized header -> internal field
HEADER_MAP = {
    "supplier's name": "supplier_name",
    "suppliers name": "supplier_name",
    "supplier's bin": "supplier_bin",
    "suppliers bin": "supplier_bin",
    "tax deposit serial no.": "treasury_challan_no",
    "tax deposit serial no": "treasury_challan_no",
    "tax deposit serial no.of bank transfer": "treasury_challan_no",
    "tax deposit date": "treasury_deposit_date",
    "concerned invoiced number": "invoice_no",
    "concerned invoice number": "invoice_no",
    "invoice/challan/bill date": "invoice_date",
    "deducted vat": "deducted_vat",
    "vat rate": "vat_rate",
    "notes": "category",
    "service (notes)": "category",
}


def _norm(v) -> str:
    return str(v or "").strip().lower()


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "").replace("%", ""))
    except ValueError:
        return None


def parse_summary(path: str, sheet_name: str = "Summary") -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[-1]]

    # find the header row: the row whose cells match the most known column labels.
    # (a loose substring search on "supplier"/"vat" would also match a title row like
    # "Sub Form for Note 24 (VAT Deducted at Source from Suppliers)")
    header_row, best_score = None, 0
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [_norm(c.value) for c in ws[r]]
        score = sum(1 for v in vals if v in HEADER_MAP)
        if score > best_score:
            header_row, best_score = r, score
    if header_row is None or best_score < 2:
        raise ValueError("Could not locate the Summary header row (expected known column labels "
                         "like \"Supplier's Name\" and \"Deducted VAT\").")

    cols = {}
    for idx, cell in enumerate(ws[header_row]):
        field = HEADER_MAP.get(_norm(cell.value))
        if field:
            cols[field] = idx

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        cells = ws[r]
        if all((c.value in (None, "")) for c in cells):
            continue
        def g(f):
            i = cols.get(f)
            return cells[i].value if i is not None and i < len(cells) else None

        rate = _to_float(g("vat_rate"))
        # rates entered as "10" mean 10%
        if rate is not None and rate > 1:
            rate = rate / 100.0
        rows.append({
            "excel_row": r,
            "supplier_name": (str(g("supplier_name")).strip() if g("supplier_name") else ""),
            "supplier_bin": (str(g("supplier_bin")).strip() if g("supplier_bin") else ""),
            "treasury_challan_no": (str(g("treasury_challan_no")).strip() if g("treasury_challan_no") else ""),
            "treasury_deposit_date": _to_date(g("treasury_deposit_date")),
            "invoice_no": (str(g("invoice_no")).strip() if g("invoice_no") else ""),
            "invoice_date": _to_date(g("invoice_date")),
            "category": (str(g("category")).strip() if g("category") else ""),
            "deducted_vat": _to_float(g("deducted_vat")),
            "vat_rate": rate,
        })
    return rows
